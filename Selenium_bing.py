from selenium import webdriver
import time
import login_data
import random
from openpyxl import load_workbook
import datetime
from msedge.selenium_tools import Edge
from msedge.selenium_tools import EdgeOptions
from colorama import Fore, Back, init


path_wortlist = 'wordlist.txt'
username_list = login_data.get_username_login()
password_list = login_data.get_password_login()
maxwait = 3
minwait = 2
shortmin = 0.2
shortmax = 1
counter_acc = 0
counter_excel = counter_acc + 1


edge_options = EdgeOptions()
edge_options.add_argument('--disable-logging')
edge_options.add_argument('log-level=3')
edge_options.add_experimental_option('excludeSwitches', ['enable-logging'])
# edge_options.add_argument('headless')
# edge_options.add_argument('disable-gpu')
edge_options.use_chromium = True
init()


def shortsleep():
    time.sleep(random.uniform(shortmin, shortmax))


def sleep():
    time.sleep(random.uniform(minwait, maxwait))


def bing_pc_search():
    driver.get('https://www.bing.com/')
    for i in range(get_pc_search()):
        words = open(path_wortlist).read().split('\n')
        bing_search_field = driver.find_element_by_id('sb_form_q')
        bing_search_field.send_keys(random.choice(words) + ' ')
        shortsleep()
        bing_search_field.submit()
        shortsleep()


def get_pc_search():
    driver.get('https://www.bing.com///rewardsapp///flyout')

    sleep()
    element = driver.find_element_by_xpath(
        '//*[@id="modern-flyout"]/div/div[5]/div/div[2]/div[1]/div/div').text.split("/")
    element1 = driver.find_element_by_xpath(
        '//*[@id="modern-flyout"]/div/div[5]/div/div[2]/div[2]/div/div').text.split("/")  # probaly wrong div
    number = (int(element[1]) - int(element[0]) +
              (int(element1[1]) - int(element1[0])))/3
    driver.get('https://www.bing.com/')
    print(str(number) + ' pc Search |', end='')
    return int(number)


def get_mobile_search():
    mobile_driver.get('https://www.bing.com///rewardsapp///flyout')

    element = mobile_driver.find_element_by_xpath(
        '//*[@id="modern-flyout"]/div/div[5]/div/div[2]/div[3]/div/div').text.split("/")
    number = (int(element[1]) - int(element[0]))/3
    mobile_driver.get('https://www.bing.com/')
    print(str(number) + ' mobil search |', end='')
    return int(number)


def bing_mobile_search():
    for i in range(get_mobile_search()):
        words = open(path_wortlist).read().split('\n')
        bing_search_mobile_field = mobile_driver.find_element_by_id(
            'sb_form_q')
        bing_search_mobile_field.send_keys(random.choice(words) + ' ')
        shortsleep()
        bing_search_mobile_field.submit()
        shortsleep()


def login_bing(counter_acc_int):
    driver.get('https://www.bing.com/?cc=de')
    time.sleep(5)
    try:
        driver.find_element_by_class_name('bnp_btn_accept').click()
       # driver.find_element_by_id('bnp_btn_accept').click()
    except:
        driver.find_element_by_id('bnp_btn_accept').click()
        print('error 23123 no cookie button found', end='')
    sleep()
    try:
        sleep()
        driver.find_element_by_id('id_s').click()
        sleep()
        driver.find_element_by_class_name('id_link_text').click()
        sleep()
        driver.find_element_by_id('id_s').click()
        sleep()
    except:
        print('error login 3', end='')
    try:
        shortsleep()
        driver.find_element_by_id('i0116').send_keys(
            username_list[counter_acc_int])
        sleep()
        driver.find_element_by_id('idSIButton9').click()
        sleep()
    except:
        print('could not find login field now try restart |', end='')
        login_bing(counter_acc)
    try:
        driver.find_element_by_id('msaTile').click()
        sleep()
    except:
        print('', end='')
    driver.find_element_by_id('i0118').send_keys(
        password_list[counter_acc_int])
    sleep()
    driver.find_element_by_id('idSIButton9').click()


def get_mobile_driver():
    mobile_emulation = {"deviceName": "Nexus 5"}
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option(
        "mobileEmulation", mobile_emulation)
    chrome_options.add_argument('--disable-logging')
    chrome_options.add_argument('log-level=3')
    chrome_options.add_experimental_option(
        'excludeSwitches', ['enable-logging'])
    return webdriver.Chrome('chromedriver.exe', options=chrome_options)


def bing_mobile_login(counter_acc_int):
    mobile_driver.get('https://www.bing.com/')
    time.sleep(4)
    try:
        mobile_driver.find_element_by_id('bnp_btn_accept').click()
    except:
        print('error 1903481729471 no cookie button found |', end='')
    sleep()
    mobile_driver.find_element_by_id('mHamburger').click()
    sleep()
    mobile_driver.find_element_by_id('hb_s').click()
    sleep()
    # dont know what is happening some time works sometimens not
    mobile_driver.find_element_by_id(
        'i0116').send_keys(username_list[counter_acc_int])
    sleep()
    mobile_driver.find_element_by_id('idSIButton9').click()
    sleep()
    try:
        mobile_driver.find_element_by_id('msaTile').click()
        sleep()
    except:
        print('', end='')
    mobile_driver.find_element_by_id(
        'i0118').send_keys(password_list[counter_acc_int])
    sleep()
    mobile_driver.find_element_by_id('idSIButton9').click()
    sleep()


def scan_daily_task():
    driver.get('https://www.bing.com///rewardsapp///flyout')
    list_len = len(driver.find_elements_by_class_name('rw-si.add'))
    print("List_len: " + str(list_len) + '|', end='')
    while list_len != 0:
        driver.find_elements_by_class_name('rw-si.add')[0].click()
        bing_do_task()
        driver.get('https://www.bing.com///rewardsapp///flyout')
        sleep()
        list_len = len(driver.find_elements_by_class_name('rw-si.add'))
        print("List_len: " + str(list_len) + '|', end='')
    sleep()
    driver.find_element_by_class_name('i-h.rw-sh.fp_row').click()
    list_len = len(driver.find_elements_by_class_name('rw-si.add'))
    print("List_len: " + str(list_len) + '|', end='')
    while list_len != 0:
        driver.find_elements_by_class_name('rw-si.add')[0].click()
        bing_do_task()
        driver.get('https://www.bing.com///rewardsapp///flyout')
        sleep()
        driver.find_element_by_class_name('i-h.rw-sh.fp_row').click()
        list_len = len(driver.find_elements_by_class_name('rw-si.add'))
        print("List_len: " + str(list_len) + '|', end='')


def bing_do_task():
    sleep()  # eddit
    sleep()
    try:
        driver.find_element_by_id('rqStartQuiz')
        quiz = True
    except:
        quiz = False
        print('', end='')
    try:
        driver.find_element_by_id('btoption0')
        survery = True
    except:
        survery = False
        print('', end='')

    if survery == True:

        driver.find_element_by_id('btoption0').click()
        sleep()

    if quiz == True:
        driver.find_element_by_id('rqStartQuiz').click()
        sleep()
        try:
            driver.find_element_by_class_name('btCardImg')
            do_quiz_3()
            # return 0
        except:
            print('', end='')
        try:
            driver.find_elements_by_class_name("bt_cardText")
            do_quiz_1()
            # return
        except:
            print('', end='')
        try:
            driver.find_element_by_id('rqAnswerOption0')
            do_quiz_2()
        except:
            print('', end='')
        try:
            # ich binn dumm denk nach laurin
            driver.find_elements_by_css_selector(
                'div.b_cards[iscorrectoption="True"]')
            do_quiz_1()
            print('do quiz 1 again beacause it failed before')
        except:
            print('', end='')


def do_quiz_1():
    list_len = len(driver.find_elements_by_css_selector(
        'div.b_cards[iscorrectoption="True"]'))
    print("List_len: " + str(list_len) + ' |', end='')
    for i in range(3):  # 3 ?
        time.sleep(7)  # hin
        for t in range(len(driver.find_elements_by_css_selector('div.b_cards[iscorrectoption="True"]'))):
            sleep()

            try:
                driver.find_elements_by_css_selector(
                    'div.b_cards[iscorrectoption="True"]')[t].click()
                sleep()  # hinzugefügt
            except:
                print('something went wrong |', end='')


def do_quiz_2():
    button_quiz = driver.find_elements_by_class_name('rq_button')
    sleep()
    for i in range(3):
        sleep()
        for i in range(len(button_quiz)):
            sleep()
            try:
                button_quiz = driver.find_elements_by_class_name('rq_button')
                button_quiz[i].click()
            except:
                print('The Cake is a lie |', end='')


def do_quiz_3():
    for d in range(10):
        sleep()
        button_quiz_3 = driver.find_elements_by_class_name(
            'btCardImg')
        sleep()
        button_quiz_3[1].click()


def get_points():
    driver.get('https://www.bing.com/')
    sleep()
    points = driver.find_element_by_id('id_rc').text
    return points


def get_excel_data(counter_excel):
    wb = load_workbook(r"C:\Users\lauri\Desktop\points.xlsx")
    points = get_points()
    wb.active = counter_excel
    ws = wb.active
    first_day = datetime.date(2021, 6, 30)
    today = datetime.date.today()
    sub_date = today - first_day
    sub_date = sub_date.days + 2
    ws.cell(row=sub_date, column=2, value=points)
    ws.cell(row=sub_date, column=1, value=today.strftime('%d.%m.%Y'))
    wb.save(r"C:\Users\lauri\Desktop\points.xlsx")


for j in range(len(username_list)):
    driver = Edge(executable_path='msedgedriver.exe', options=edge_options)
    driver.get('https://bing.com')
    login_bing(counter_acc)
    scan_daily_task()
    print('daily task done |', end='')
    bing_pc_search()
    print('bing search done |', end='')
    get_excel_data(counter_excel)
    driver.quit()

    mobile_driver = get_mobile_driver()
    bing_mobile_login(counter_acc)
    bing_mobile_search()
    print('mobile search done |',)
    mobile_driver.quit()
    if counter_acc == 0:
        print('\033[32m' + 'Laurin Done')
        print('\033[39m')
    elif counter_acc == 1:
        print('\033[32m' + 'Tim Done')
        print('\033[39m')
    elif counter_acc == 2:
        print('\033[32m' + 'Laurin_2 Done')
        print('\033[39m')
    elif counter_acc == 3:
        print('\033[32m' + 'Marvin Done')
        print('\033[39m')
    counter_excel = counter_excel + 1
    counter_acc = counter_acc + 1
